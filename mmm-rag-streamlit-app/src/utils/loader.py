from openpyxl import load_workbook

def load_data(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    data = []
    
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    return data

def load_excel_data():
    data_path = "data/MMM Dummy Data.xlsx"
    return load_data(data_path)